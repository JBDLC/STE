# main.py

from fastapi import FastAPI, Form
from openpyxl import Workbook, load_workbook
import os

app = FastAPI()

EXCEL_FILE = "data.xlsx"

# Crée un fichier Excel si inexistant
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Nom", "Prénom", "Email", "Question1", "Question2"])
    wb.save(EXCEL_FILE)

@app.post("/submit")
async def submit_form(nom: str = Form(...), prenom: str = Form(...),
                      email: str = Form(...), question1: str = Form(...),
                      question2: str = Form(...)):
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([nom, prenom, email, question1, question2])
    wb.save(EXCEL_FILE)
    
    return {"message": "Données enregistrées avec succès."}
